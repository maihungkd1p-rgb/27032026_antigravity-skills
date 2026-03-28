"""
Excel Professional Template
===========================
Template chuẩn tạo file Excel với định dạng doanh nghiệp.

Sử dụng:
    from excel_template import ExcelBuilder
    
    builder = ExcelBuilder("Báo cáo Q1 2026")
    builder.add_header_row(["STT", "Nội dung", "Giá trị"])
    builder.add_data_rows(data)
    builder.save("output.xlsx")
"""

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from typing import List, Dict, Optional, Tuple, Any


class ExcelStyles:
    """Định nghĩa styles chuẩn cho Excel"""
    
    # === FONTS ===
    HEADER_FONT = Font(bold=True, size=11, color="FFFFFF")
    TITLE_FONT = Font(bold=True, size=14)
    GROUP_FONT = Font(bold=True, size=11)
    BOLD_FONT = Font(bold=True)
    NORMAL_FONT = Font(size=11)
    
    # === FILLS ===
    HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    GROUP_FILL = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
    ALT_ROW_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    
    # Accent colors
    GREEN_FILL = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
    ORANGE_FILL = PatternFill(start_color="ED7D31", end_color="ED7D31", fill_type="solid")
    RED_FILL = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
    YELLOW_FILL = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
    
    # === BORDERS ===
    THIN_BORDER = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    THICK_BOTTOM = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='medium')
    )
    
    # === ALIGNMENTS ===
    CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
    LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
    RIGHT = Alignment(horizontal="right", vertical="center", wrap_text=True)


class ExcelBuilder:
    """Builder class tạo file Excel chuyên nghiệp"""
    
    def __init__(self, title: str = ""):
        """
        Khởi tạo ExcelBuilder
        
        Args:
            title: Tiêu đề file (hiển thị ở dòng 1)
        """
        self.wb = Workbook()
        self.ws = self.wb.active
        self.title = title
        self.current_row = 1
        self.styles = ExcelStyles()
        
        if title:
            self._add_title(title)
    
    def _add_title(self, title: str, end_col: int = 10):
        """Thêm tiêu đề merge cells"""
        self.ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=end_col)
        self.ws['A1'] = title
        self.ws['A1'].font = self.styles.TITLE_FONT
        self.ws['A1'].alignment = self.styles.CENTER
        self.current_row = 2
    
    def add_header_row(self, headers: List[str], row: Optional[int] = None):
        """
        Thêm dòng header
        
        Args:
            headers: Danh sách tên cột
            row: Số dòng (mặc định: current_row)
        """
        if row is None:
            row = self.current_row
        
        for col, header in enumerate(headers, 1):
            cell = self.ws.cell(row=row, column=col, value=header)
            cell.font = self.styles.HEADER_FONT
            cell.fill = self.styles.HEADER_FILL
            cell.alignment = self.styles.CENTER
            cell.border = self.styles.THIN_BORDER
        
        self.current_row = row + 1
        return self
    
    def add_group_row(self, code: str, name: str, end_col: int, row: Optional[int] = None):
        """
        Thêm dòng nhóm (với merge cells)
        
        Args:
            code: Mã nhóm (cột 1)
            name: Tên nhóm (merge từ cột 2)
            end_col: Cột cuối cùng
            row: Số dòng (mặc định: current_row)
        """
        if row is None:
            row = self.current_row
        
        # Cột 1: Mã nhóm
        cell1 = self.ws.cell(row=row, column=1, value=code)
        cell1.font = self.styles.GROUP_FONT
        cell1.fill = self.styles.GROUP_FILL
        cell1.alignment = self.styles.CENTER
        cell1.border = self.styles.THIN_BORDER
        
        # Merge cột 2 đến end_col
        self.ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=end_col)
        cell2 = self.ws.cell(row=row, column=2, value=name)
        cell2.font = self.styles.GROUP_FONT
        cell2.fill = self.styles.GROUP_FILL
        cell2.alignment = self.styles.LEFT
        
        # Border cho tất cả cells
        for col in range(1, end_col + 1):
            self.ws.cell(row=row, column=col).border = self.styles.THIN_BORDER
        
        self.current_row = row + 1
        return self
    
    def add_data_row(self, data: List[Any], row: Optional[int] = None, 
                     center_from_col: int = 3):
        """
        Thêm dòng dữ liệu
        
        Args:
            data: Danh sách giá trị
            row: Số dòng (mặc định: current_row)
            center_from_col: Căn giữa từ cột này trở đi
        """
        if row is None:
            row = self.current_row
        
        for col, value in enumerate(data, 1):
            cell = self.ws.cell(row=row, column=col, value=value)
            cell.border = self.styles.THIN_BORDER
            cell.alignment = self.styles.CENTER if col >= center_from_col else self.styles.LEFT
        
        self.current_row = row + 1
        return self
    
    def add_data_rows(self, rows: List[List[Any]], center_from_col: int = 3):
        """Thêm nhiều dòng dữ liệu"""
        for row_data in rows:
            self.add_data_row(row_data, center_from_col=center_from_col)
        return self
    
    def add_timeline_row(self, stt: str, task: str, months: List[int], 
                         marker: str = "●", row: Optional[int] = None):
        """
        Thêm dòng timeline (kế hoạch theo tháng)
        
        Args:
            stt: Số thứ tự
            task: Tên công việc
            months: Danh sách tháng thực hiện (1-12)
            marker: Ký hiệu đánh dấu
            row: Số dòng
        """
        if row is None:
            row = self.current_row
        
        # STT và Task
        self.ws.cell(row=row, column=1, value=stt).border = self.styles.THIN_BORDER
        self.ws.cell(row=row, column=1).alignment = self.styles.CENTER
        self.ws.cell(row=row, column=2, value=task).border = self.styles.THIN_BORDER
        self.ws.cell(row=row, column=2).alignment = self.styles.LEFT
        
        # Months (cột 3-14)
        for col in range(3, 15):
            month = col - 2
            cell = self.ws.cell(row=row, column=col)
            if month in months:
                cell.value = marker
            cell.alignment = self.styles.CENTER
            cell.border = self.styles.THIN_BORDER
        
        self.current_row = row + 1
        return self
    
    def set_column_widths(self, widths: Dict[int, int]):
        """
        Đặt độ rộng cột
        
        Args:
            widths: Dict {column_number: width}
        """
        for col, width in widths.items():
            self.ws.column_dimensions[get_column_letter(col)].width = width
        return self
    
    def auto_column_widths(self, min_width: int = 8, max_width: int = 50):
        """Tự động điều chỉnh độ rộng cột"""
        for column_cells in self.ws.columns:
            length = max(len(str(cell.value or "")) for cell in column_cells)
            adjusted_width = min(max(length + 2, min_width), max_width)
            col_letter = get_column_letter(column_cells[0].column)
            self.ws.column_dimensions[col_letter].width = adjusted_width
        return self
    
    def freeze_panes(self, cell: str = "A3"):
        """Đóng băng panes (cố định header)"""
        self.ws.freeze_panes = cell
        return self
    
    def add_auto_filter(self):
        """Thêm auto filter cho toàn bộ dữ liệu"""
        self.ws.auto_filter.ref = self.ws.dimensions
        return self
    
    def create_sheet(self, name: str, title: str = "") -> 'ExcelBuilder':
        """
        Tạo sheet mới và trả về builder mới
        
        Args:
            name: Tên sheet (max 31 ký tự)
            title: Tiêu đề sheet
        """
        self.ws = self.wb.create_sheet(name[:31])
        self.current_row = 1
        if title:
            self._add_title(title)
        return self
    
    def save(self, filepath: str):
        """Lưu file Excel"""
        self.wb.save(filepath)
        print(f"✅ Đã lưu: {filepath}")
        return filepath


# === UTILITY FUNCTIONS ===

def create_timeline_excel(
    title: str,
    data: List[Tuple[str, str, List[int]]],  # [(stt, task, months), ...]
    groups: Dict[str, str] = None,  # {"A": "Nhóm A", ...}
    output_path: str = "output.xlsx"
) -> str:
    """
    Tạo file Excel timeline/kế hoạch
    
    Args:
        title: Tiêu đề
        data: Danh sách (stt, task, months) hoặc (group_code, group_name, None)
        groups: Dict định nghĩa group codes (optional)
        output_path: Đường dẫn file xuất
    
    Returns:
        Đường dẫn file đã tạo
    """
    builder = ExcelBuilder(title)
    
    # Header
    headers = ['STT', 'Công việc'] + [f'T{i}' for i in range(1, 13)]
    builder.add_header_row(headers)
    
    # Data
    for stt, task, months in data:
        if months is None:
            builder.add_group_row(stt, task, 14)
        else:
            builder.add_timeline_row(stt, task, months)
    
    # Column widths
    builder.set_column_widths({1: 6, 2: 55})
    for col in range(3, 15):
        builder.ws.column_dimensions[get_column_letter(col)].width = 5
    
    builder.freeze_panes("C3")
    builder.save(output_path)
    
    return output_path


def create_summary_excel(
    title: str,
    data: List[Tuple[str, Any]],  # [(label, value), ...]
    output_path: str = "summary.xlsx",
    total_label: str = "Tổng cộng"
) -> str:
    """
    Tạo file Excel bảng tổng hợp
    
    Args:
        title: Tiêu đề
        data: Danh sách (label, value)
        output_path: Đường dẫn file xuất
        total_label: Nhãn dòng tổng (sẽ được bold)
    """
    builder = ExcelBuilder(title)
    builder.add_header_row(["Hạng mục", "Giá trị"])
    
    for label, value in data:
        row = builder.current_row
        builder.ws.cell(row=row, column=1, value=label)
        builder.ws.cell(row=row, column=2, value=value)
        builder.ws.cell(row=row, column=1).border = builder.styles.THIN_BORDER
        builder.ws.cell(row=row, column=2).border = builder.styles.THIN_BORDER
        builder.ws.cell(row=row, column=2).alignment = builder.styles.CENTER
        
        if label == total_label:
            builder.ws.cell(row=row, column=1).font = builder.styles.BOLD_FONT
            builder.ws.cell(row=row, column=2).font = builder.styles.BOLD_FONT
        
        builder.current_row += 1
    
    builder.set_column_widths({1: 30, 2: 20})
    builder.save(output_path)
    
    return output_path


# === MAIN (Demo) ===

if __name__ == "__main__":
    # Demo: Tạo file timeline
    demo_data = [
        ("A", "Nhóm công việc A", None),
        ("1.1", "Nhiệm vụ 1", [1, 2, 3]),
        ("1.2", "Nhiệm vụ 2", [4, 5, 6, 7]),
        ("B", "Nhóm công việc B", None),
        ("2.1", "Nhiệm vụ 3", [6, 7, 8, 9, 10]),
        ("2.2", "Nhiệm vụ 4", [10, 11, 12]),
    ]
    
    create_timeline_excel(
        "KẾ HOẠCH DEMO 2026",
        demo_data,
        output_path="demo_timeline.xlsx"
    )
    
    # Demo: Tạo bảng tổng hợp
    summary = [
        ("Hạng mục A", 100),
        ("Hạng mục B", 200),
        ("Hạng mục C", 150),
        ("Tổng cộng", 450),
    ]
    
    create_summary_excel(
        "BẢNG TỔNG HỢP DEMO",
        summary,
        output_path="demo_summary.xlsx"
    )
