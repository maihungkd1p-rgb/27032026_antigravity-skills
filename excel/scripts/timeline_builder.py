"""
Timeline Excel Template - Pattern 4
Xuất bảng kế hoạch timeline với đánh dấu x và fill màu nền

Cách sử dụng:
1. Import các hàm và constants
2. Chuẩn bị data theo format:
   - {"is_group": True, "name": "Tên nhóm"}
   - {"stt": "1.1", "name": "Tên công việc", "months": [0,1,1,1,0,0,0,0,0,0,0,0]}
3. Gọi create_timeline_sheet(ws, data)
"""
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

# ===== STYLES =====
HEADER_FONT = Font(bold=True, size=11, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
GROUP_FILL = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
GROUP_FONT = Font(bold=True, size=11)
MARK_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)
CENTER = Alignment(horizontal='center', vertical='center')
LEFT = Alignment(horizontal='left', vertical='center')


def create_timeline_sheet(ws, data, title=None):
    """
    Tạo sheet timeline với format chuẩn
    
    Args:
        ws: worksheet object
        data: list of dicts với format:
            - {"is_group": True, "name": "Tên nhóm"}
            - {"stt": "1.1", "name": "Tên CV", "months": [0,1,1,1,0,0,0,0,0,0,0,0]}
        title: optional title row
    """
    start_row = 1
    
    # Title row (optional)
    if title:
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=14)
        ws.cell(row=1, column=1, value=title)
        ws.cell(row=1, column=1).font = Font(bold=True, size=14)
        ws.cell(row=1, column=1).alignment = CENTER
        start_row = 2
    
    # Header row
    headers = ["STT", "Công việc", "T1", "T2", "T3", "T4", "T5", "T6", "T7", "T8", "T9", "T10", "T11", "T12"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=start_row, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = THIN_BORDER
    
    # Data rows
    row = start_row + 1
    for item in data:
        if item.get("is_group"):
            # Group header row
            cell = ws.cell(row=row, column=1, value=item["name"])
            cell.font = GROUP_FONT
            cell.fill = GROUP_FILL
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=14)
            for col in range(1, 15):
                ws.cell(row=row, column=col).border = THIN_BORDER
                ws.cell(row=row, column=col).fill = GROUP_FILL
        else:
            # Regular row
            ws.cell(row=row, column=1, value=item["stt"]).alignment = CENTER
            ws.cell(row=row, column=2, value=item["name"]).alignment = LEFT
            
            # Month marks
            for i, month in enumerate(item.get("months", []), 3):
                cell = ws.cell(row=row, column=i)
                if month:
                    cell.value = "x"
                    cell.fill = MARK_FILL
                cell.alignment = CENTER
            
            # Apply borders
            for col in range(1, 15):
                ws.cell(row=row, column=col).border = THIN_BORDER
        
        row += 1
    
    # Column widths
    ws.column_dimensions['A'].width = 6
    ws.column_dimensions['B'].width = 55
    for col in range(3, 15):
        ws.column_dimensions[get_column_letter(col)].width = 5
    
    # Freeze panes
    ws.freeze_panes = f'C{start_row + 1}'
    
    return row  # Return next available row


def create_timeline_workbook(sheets_data: list, output_path: str):
    """
    Tạo workbook với nhiều sheet timeline
    
    Args:
        sheets_data: list of tuples (sheet_name, data)
        output_path: đường dẫn file output
    
    Example:
        sheets_data = [
            ("Sheet1", data1),
            ("Sheet2", data2),
        ]
    """
    wb = Workbook()
    
    for i, (sheet_name, data) in enumerate(sheets_data):
        if i == 0:
            ws = wb.active
            ws.title = sheet_name[:31]
        else:
            ws = wb.create_sheet(sheet_name[:31])
        
        create_timeline_sheet(ws, data)
    
    wb.save(output_path)
    print(f"Đã lưu file: {output_path}")
    return wb


# ===== USAGE EXAMPLE =====
if __name__ == "__main__":
    # Example data
    sample_data = [
        {"is_group": True, "name": "A. Nhóm công việc 1"},
        {"stt": "1.1", "name": "Công việc mẫu 1", "months": [0,1,1,1,0,0,0,0,0,0,0,0]},
        {"stt": "1.2", "name": "Công việc mẫu 2", "months": [0,0,0,1,1,1,1,0,0,0,0,0]},
        {"is_group": True, "name": "B. Nhóm công việc 2"},
        {"stt": "2.1", "name": "Công việc mẫu 3", "months": [0,0,0,0,0,1,1,1,1,1,0,0]},
    ]
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Ke hoach"
    create_timeline_sheet(ws, sample_data)
    wb.save("timeline_example.xlsx")
    print("Created timeline_example.xlsx")
